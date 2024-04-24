import openpyxl
import os
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm

# Function to check if a file exists and return an incremented name
def get_unique_filename(filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    while os.path.exists(filename):
        filename = f"{base}_{counter}{ext}"
        counter += 1
    return filename

# Path to the Excel file containing the URLs
excel_file_path = "C:\\Users\\MSI Thin\\Documents\\chrry\\script\\script-py\\excels\\2020A_1.xlsx"

# Define the index of the column containing the URLs (assuming it's the second column)
url_column_index = 2
url_column_status_posttest = 7
url_column_status_pretest = 8
url_column_pretest_order = 9  # New column for pre test order
url_column_post_test_order = 10  # New column for post test order

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Get the active worksheet
worksheet = workbook.active

# Function to process a single URL
def process_url(url):
    try:
        response = requests.get(url)
        print("Opened URL:", url)

        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find all tables in the page
        tables = soup.find_all('table')
        
        # Access text from the last table
        last_table_text = tables[-1].get_text()

        # Check if "Post Test" or "Ujian" is present in the last table text
        if "Post Test" in last_table_text or "Ujian" in last_table_text:
            # Check if "Post Test" or "Ujian" is at the last position in the text
            if last_table_text.strip().endswith(("Post Test", "Ujian")):
                print("Post Test/Ujian ditemukan di urutan terakhir pada tabel")
                post_test_result = "OK"
            else:
                print("Error: Post Test/Ujian tidak ditemukan di urutan terakhir pada tabel")
                post_test_result = "ERROR"
        else:
            print("Post Test/Ujian tidak ditemukan di tabel")
            post_test_result = "ERROR"

        # Check if tbody with id "encounterList" exists
        tbody_encounter_list = soup.find('tbody', id='encounterList')
        
        if tbody_encounter_list is None or not tbody_encounter_list.find_all(['tr', 'td']):
            return post_test_result, "Tidak Ada", None, None  # Return "Tidak Ada" for pre test status and None for orders
        
        # Check for Pre Test and Post Test orders
        pre_test_order = None
        post_test_order = None
        for idx, row in enumerate(tbody_encounter_list.find_all('tr'), start=1):
            if "Pre Test" in row.text:
                pre_test_order = idx
            if "Post Test" in row.text or "Ujian" in row.text:
                post_test_order = idx
        
        pre_test_status = "Ada" if pre_test_order else "Tidak Ada"
        
        return post_test_result, pre_test_status, pre_test_order, post_test_order

    except Exception as e:
        print("Error processing URL:", url, "-", e)
        return "ERROR", "ERROR", None, None


# List to store results
results = []

# Iterate through each row in Excel data
urls = [row[url_column_index - 1] for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True) if row[url_column_index - 1]]

# Define the total number of URLs for tqdm
total_urls = len(urls)

# Function to process URLs and update progress bar
def process_urls_with_progress(url):
    post_test_result, pre_test_result, pre_test_order, post_test_order = process_url(url)
    results.append((post_test_result, pre_test_result, pre_test_order, post_test_order))
    pbar.update(1)  # Update progress bar

# Set up tqdm progress bar
with tqdm(total=total_urls, desc="Processing URLs", unit="URL") as pbar:
    # Process URLs using ThreadPoolExecutor
    with ThreadPoolExecutor() as executor:
        for url in urls:
            executor.submit(process_urls_with_progress, url)

# Insert new columns for status post test, status pre test, and pre test order
worksheet.insert_cols(url_column_status_posttest, 4)

# Write titles for the new columns
worksheet.cell(row=1, column=url_column_status_posttest, value="Status Post Test/Ujian")
worksheet.cell(row=1, column=url_column_status_pretest, value="Status Pre Test")
worksheet.cell(row=1, column=url_column_pretest_order, value="Urutan Pretest")  # Column for pre test order
worksheet.cell(row=1, column=url_column_post_test_order, value="Urutan Post Test/Ujian")  # Column for post test order

# Update the status columns with results
for row_num, (post_test_result, pre_test_result, pre_test_order, post_test_order) in zip(range(2, len(urls) + 2), results):
    worksheet.cell(row=row_num, column=url_column_status_posttest, value=post_test_result)
    worksheet.cell(row=row_num, column=url_column_status_pretest, value=pre_test_result)
    worksheet.cell(row=row_num, column=url_column_pretest_order, value=pre_test_order if pre_test_order is not None else "Error")
    worksheet.cell(row=row_num, column=url_column_post_test_order, value=post_test_order if post_test_order is not None else "Error")
    print(f"Updated 'Status Post Test/Ujian' column with: {post_test_result}")
    print(f"Updated 'Status Pre Test' column with: {pre_test_result}")
    print(f"Updated 'Urutan Pretest' column with: {pre_test_order}")
    print(f"Updated 'Urutan Post Test/Ujian' column with: {post_test_order}")

# Generate a unique filename for export
export_filename = get_unique_filename("C:\\Users\\MSI Thin\\Documents\\chrry\\script\\script-py\\excels\\TES.xlsx")

# Save the workbook with the unique filename
workbook.save(export_filename)
print("Exported data to:", export_filename)

# Close the workbook
workbook.close()
