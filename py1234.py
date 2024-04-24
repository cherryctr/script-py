import openpyxl
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import logging

# Setup logging
logging.basicConfig(filename='process_errors.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def setup_session():
    """Setup a requests session with retry logic."""
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"]  # Corrected parameter
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

def fetch_and_process(session, url, row_num, url_column_postest, url_column_pretest):
    """Fetch URL content and extract status of 'Post Test' and 'Pre Test'."""
    try:
        response = session.get(url, timeout=30)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            tbody_encounter_list = soup.find('tbody', id="encounterList")

            post_status = "Belum Post Test"
            pre_status = "Belum Pre Test"

            if tbody_encounter_list and tbody_encounter_list.find_all(['tr', 'td']):
                if "Post Test" in soup.text:
                    post_status = "Sudah Post Test"
                if "Pre Test" in soup.text:
                    pre_status = "Sudah Pre Test"

            return row_num, post_status, pre_status
        else:
            logging.error(f"Failed to retrieve: {url} - Status Code: {response.status_code}")
            return row_num, "Failed to Retrieve", "Failed to Retrieve"
    except Exception as e:
        logging.error(f"Error processing URL: {url} - {e}")
        return row_num, "Error", "Error"

session = setup_session()

# Load workbook and worksheet
excel_file_path = "C:\\Users\\MSI Thin\\Documents\\chrry\\script\\script-py\\excels\\2020A_1.xlsx"
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active

# Determine necessary column indices
materi_column_index = find_column_index(worksheet, "Materi")
url_column_index = 2
url_column_postest = find_column_index(worksheet, "Status Postest", materi_column_index + 1)
url_column_pretest = find_column_index(worksheet, "Status Pre Test", materi_column_index + 2)

# Prepare URL list with associated row numbers
urls = [(row[url_column_index - 1], row_num) for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_col=url_column_index + 1, values_only=True), start=2) if row[url_column_index - 1]]

# Process URLs in parallel
with ThreadPoolExecutor(max_workers=10) as executor:
    future_to_url = {executor.submit(fetch_and_process, session, url, row_num, url_column_postest, url_column_pretest): url for url, row_num in urls}
    for future in tqdm(as_completed(future_to_url), total=len(urls), desc="Processing URLs"):
        row_num, post_status, pre_status = future.result()
        worksheet.cell(row=row_num, column=url_column_postest, value=post_status)
        worksheet.cell(row=row_num, column=url_column_pretest, value=pre_status)

# Save and close workbook
excel_export_path = excel_file_path.replace('.xlsx', '_BELUM_LULUS_RESULT_FIX.xlsx')
try:
    workbook.save(excel_export_path)
    logging.info("Exported data to: " + excel_export_path)
except PermissionError as e:
    logging.error("PermissionError: " + str(e))
finally:
    workbook.close()
