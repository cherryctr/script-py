import openpyxl
import requests
from bs4 import BeautifulSoup

# Path to the Excel file containing the URLs
excel_file_path = "C:\\Users\\Admin\\Documents\\pysc\\excels\\DATA_2020A.xlsx"

# Define the index of the column containing the URLs (assuming it's the second column)
url_column_index = 2
url_column_catatan = 3

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Get the active worksheet
worksheet = workbook.active

# Iterate through each row in Excel data
for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_col=2, values_only=True), start=2):  # Start from the second row
    url = row[url_column_index - 1]  # Adjust index to 0-based

    if not url:
        continue

    try:
        response = requests.get(url)
        print("Opened URL:", url)

        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Check if tbody with id "encounterList" exists
        tbody_encounter_list = soup.find('tbody', id='encounterList')
        
        if tbody_encounter_list is None or not tbody_encounter_list.find_all(['tr', 'td']):
            worksheet.cell(row=row_num, column=url_column_catatan, value="Belum Post Test")
            print("Updated 'Catatan' column with: Belum Post Test")
            continue
        
        # Check for Post Test existence
        if "Post Test" in response.text:
            worksheet.cell(row=row_num, column=url_column_catatan, value="Sudah Post Test")
            print("Updated 'Catatan' column with: Sudah Post Test")
        else:
            worksheet.cell(row=row_num, column=url_column_catatan, value="Belum Post Test")
            print("Updated 'Catatan' column with: Belum Post Test")

             # Check for Pre Test existence
        if "Pre Test" in response.text:
            worksheet.cell(row=row_num, column=url_column_catatan, value="Sudah Pre Test")
            print("Updated 'Catatan' column with: Sudah Pre Test")
        else:
            worksheet.cell(row=row_num, column=url_column_catatan, value="Belum Pre Test")
            print("Updated 'Catatan' column with: Belum Pre Test")

         # Check for Pre Test existence and position
        pre_test_exists = False
        pre_test_position = None
        for idx, tr in enumerate(tbody_encounter_list.find_all('tr')):
            if 'Pre Test' in tr.text:
                pre_test_exists = True
                if idx == 0:
                    pre_test_position = 'Paling Atas'
                break
        
            # Check for Post Test existence and position
            post_test_exists = False
            post_test_position = None
            for idx, tr in enumerate(reversed(list(tbody_encounter_list.find_all('tr')))):
                if 'Post Test' in tr.text:
                    post_test_exists = True
                    if idx == 0:
                        post_test_position = 'Paling Bawah'
                    break

            # Update Catatan column based on conditions
            if pre_test_exists:
                if pre_test_position == 'Paling Atas':
                    worksheet.cell(row=row_num, column=url_column_catatan, value="OKE")
                else:
                    worksheet.cell(row=row_num, column=url_column_catatan, value="Error: Pre Test Posisinya Tidak di Paling Atas pada Table")
            else:
                worksheet.cell(row=row_num, column=url_column_catatan, value="Error: Pre Test Tidak Ada pada Table")
            
            if post_test_exists:
                if post_test_position == 'Paling Bawah':
                    worksheet.cell(row=row_num, column=url_column_catatan, value="OKE")
                else:
                    worksheet.cell(row=row_num, column=url_column_catatan, value="Error: Post Test Tidak Sesuai Posisi")
            else:
                worksheet.cell(row=row_num, column=url_column_catatan, value="Error: Post Test Tidak Ada pada Table")

    except Exception as e:
        print("Error processing URL:", url, "-", e)

# Save the workbook
excel_export_path = "C:\\Users\\Admin\\Documents\\pysc\\excels\\2020A_BELUM_LULUS_result.xlsx"
workbook.save(excel_export_path)
print("Exported data to:", excel_export_path)

workbook.close()
